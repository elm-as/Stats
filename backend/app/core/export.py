"""
Module d'export multi-format : Excel, JSON structure, HTML autonome.
"""

from __future__ import annotations

import json
import math
from html import escape
from typing import Any

import pandas as pd


def export_excel(output_path: str, payload: dict) -> str:
    """Exporte les resultats dans un classeur Excel multi-onglets."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    numeric_stats, categorical_stats = _descriptive_frames(payload)
    used_sheet_names: set[str] = set()
    tables = [
        ("Resume", _summary_frame(payload)),
        ("Apercu", _preview_frame(payload)),
        ("Dictionnaire", _dictionary_frame(payload)),
        ("Stats Numeriques", numeric_stats),
        ("Stats Categorielles", categorical_stats),
        ("Corr Pearson", _correlation_matrix_frame(payload, "pearson")),
        ("Corr Spearman", _correlation_matrix_frame(payload, "spearman")),
        ("Corr Significatives", _significant_correlations_frame(payload)),
        ("VIF", _vif_frame(payload)),
        ("Tests", _tests_frame(payload)),
        ("Modelisation", _model_ranking_frame(payload)),
        ("Importance Vars", _feature_importance_frame(payload)),
        ("SHAP Global", _shap_frame(payload)),
        ("TS Resume", _timeseries_summary_frame(payload)),
        ("TS Previsions", _timeseries_forecast_frame(payload)),
        ("MTS Resume", _multivariate_summary_frame(payload)),
        ("Granger", _granger_frame(payload)),
        ("Johansen", _johansen_frame(payload)),
        ("MTS Previsions", _multivariate_forecast_frame(payload)),
        ("ACP Variance", _pca_variance_frame(payload)),
        ("ACP Loadings", _pca_loadings_frame(payload)),
        ("AFC Lignes", _ca_coords_frame(payload, axis="row")),
        ("AFC Colonnes", _ca_coords_frame(payload, axis="col")),
        ("ACM Modalites", _mca_modalities_frame(payload)),
        ("Nettoyage", _log_frame(payload.get("cleaning_log"))),
        ("Transformations", _log_frame(payload.get("transform_logs"))),
        ("Versions", _versions_frame(payload)),
        ("Historique", _history_frame(payload)),
        ("Audit", _audit_frame(payload)),
    ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for base_name, frame in tables:
            if frame is None or frame.empty:
                continue
            sheet_name = _safe_sheet_name(base_name, used_sheet_names)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        header_fill = PatternFill(start_color="16324F", end_color="16324F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin", color="D5DCE5"),
            right=Side(style="thin", color="D5DCE5"),
            top=Side(style="thin", color="D5DCE5"),
            bottom=Side(style="thin", color="D5DCE5"),
        )

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.sheet_view.showGridLines = False
            if worksheet.max_row >= 1:
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
            if worksheet.max_row > 1:
                worksheet.auto_filter.ref = worksheet.dimensions
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = thin_border
            for col_idx, col_cells in enumerate(worksheet.columns, 1):
                max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 12), 60)

    return output_path


def export_json(output_path: str, payload: dict) -> str:
    """Exporte tous les resultats dans un JSON structure."""
    clean_payload = _sanitize_for_json(payload)
    with open(output_path, "w", encoding="utf-8") as handle:
        json.dump(clean_payload, handle, ensure_ascii=False, indent=2)
    return output_path


def export_html(output_path: str, payload: dict) -> str:
    """Genere un rapport HTML autonome et riche en sections."""
    metadata = payload.get("metadata", {})
    dataset = payload.get("dataset", {})
    title = metadata.get("title", "Export")
    organization = metadata.get("organization") or "OpenStats — Elmas Labs"
    generated_at = metadata.get("generated_at", "")

    numeric_stats, categorical_stats = _descriptive_frames(payload)
    sections: list[tuple[str, str]] = []

    sections.append(("Jeu de donnees", _render_summary_cards(payload)))

    preview = _preview_frame(payload)
    if not preview.empty:
        sections.append(("Apercu", _render_dataframe(preview, max_rows=10)))

    dictionary = _dictionary_frame(payload)
    if not dictionary.empty:
        sections.append(("Dictionnaire", _render_dataframe(dictionary, max_rows=25)))

    if not numeric_stats.empty:
        sections.append(("Statistiques Numeriques", _render_dataframe(numeric_stats, max_rows=30)))

    if not categorical_stats.empty:
        sections.append(("Statistiques Categorielles", _render_dataframe(categorical_stats, max_rows=30)))

    significant_corr = _significant_correlations_frame(payload)
    if not significant_corr.empty:
        sections.append(("Correlations Significatives", _render_dataframe(significant_corr, max_rows=40)))

    vif = _vif_frame(payload)
    if not vif.empty:
        sections.append(("Colinearite", _render_dataframe(vif, max_rows=40)))

    tests = _tests_frame(payload)
    if not tests.empty:
        sections.append(("Tests d Hypotheses", _render_dataframe(tests, max_rows=40)))

    modeling = _model_ranking_frame(payload)
    if not modeling.empty:
        sections.append(("Modelisation", _render_dataframe(modeling, max_rows=25)))

    feature_importance = _feature_importance_frame(payload)
    if not feature_importance.empty:
        sections.append(("Importance des Variables", _render_dataframe(feature_importance, max_rows=40)))

    shap = _shap_frame(payload)
    if not shap.empty:
        sections.append(("SHAP Global", _render_dataframe(shap, max_rows=30)))

    ts_summary = _timeseries_summary_frame(payload)
    if not ts_summary.empty:
        sections.append(("Series Temporelles", _render_dataframe(ts_summary, max_rows=10)))

    ts_forecast = _timeseries_forecast_frame(payload)
    if not ts_forecast.empty:
        sections.append(("Previsions Temporelles", _render_dataframe(ts_forecast, max_rows=25)))

    mts_summary = _multivariate_summary_frame(payload)
    if not mts_summary.empty:
        sections.append(("Series Temporelles Multivariees", _render_dataframe(mts_summary, max_rows=10)))

    granger = _granger_frame(payload)
    if not granger.empty:
        sections.append(("Causalite de Granger", _render_dataframe(granger, max_rows=40)))

    johansen = _johansen_frame(payload)
    if not johansen.empty:
        sections.append(("Test de Johansen", _render_dataframe(johansen, max_rows=20)))

    pca_variance = _pca_variance_frame(payload)
    if not pca_variance.empty:
        sections.append(("ACP", _render_dataframe(pca_variance, max_rows=20)))

    pca_loadings = _pca_loadings_frame(payload)
    if not pca_loadings.empty:
        sections.append(("Loadings ACP", _render_dataframe(pca_loadings, max_rows=30)))

    ca_rows = _ca_coords_frame(payload, axis="row")
    if not ca_rows.empty:
        sections.append(("AFC Lignes", _render_dataframe(ca_rows, max_rows=30)))

    ca_cols = _ca_coords_frame(payload, axis="col")
    if not ca_cols.empty:
        sections.append(("AFC Colonnes", _render_dataframe(ca_cols, max_rows=30)))

    mca = _mca_modalities_frame(payload)
    if not mca.empty:
        sections.append(("ACM Modalites", _render_dataframe(mca, max_rows=30)))

    cleaning_log = _log_frame(payload.get("cleaning_log"))
    if not cleaning_log.empty:
        sections.append(("Nettoyage", _render_dataframe(cleaning_log, max_rows=40)))

    transform_logs = _log_frame(payload.get("transform_logs"))
    if not transform_logs.empty:
        sections.append(("Transformations", _render_dataframe(transform_logs, max_rows=40)))

    versions = _versions_frame(payload)
    if not versions.empty:
        sections.append(("Versions", _render_dataframe(versions, max_rows=25)))

    history = _history_frame(payload)
    if not history.empty:
        sections.append(("Historique", _render_dataframe(history, max_rows=25)))

    audit = _audit_frame(payload)
    if not audit.empty:
        sections.append(("Audit", _render_dataframe(audit, max_rows=25)))

    nav_links = "".join(
        f'<a href="#sec-{_slugify(section_title)}">{escape(section_title)}</a>'
        for section_title, _ in sections
    )
    rendered_sections = "".join(
        f'''
        <section id="sec-{_slugify(section_title)}" class="panel">
          <div class="panel-head">
            <h2>{escape(section_title)}</h2>
          </div>
          {section_html}
        </section>
        '''
        for section_title, section_html in sections
    )

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{escape(title)}</title>
  <style>
    :root {{
      --bg: #f5f1e8;
      --ink: #1f2933;
      --muted: #5f6c7b;
      --panel: rgba(255,255,255,0.92);
      --panel-border: rgba(20, 42, 73, 0.08);
      --accent: #103d60;
      --line: #d9e1e8;
      --shadow: 0 18px 48px rgba(16, 61, 96, 0.12);
    }}
    * {{ box-sizing: border-box; }}
    html {{ scroll-behavior: smooth; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", "Helvetica Neue", sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at top left, rgba(196, 107, 61, 0.18), transparent 26%),
        radial-gradient(circle at top right, rgba(16, 61, 96, 0.12), transparent 30%),
        linear-gradient(180deg, #faf7f2 0%, var(--bg) 100%);
      line-height: 1.55;
    }}
    .shell {{ max-width: 1320px; margin: 0 auto; padding: 32px 24px 56px; }}
    .hero {{
      background: linear-gradient(135deg, rgba(16, 61, 96, 0.96), rgba(17, 92, 122, 0.92));
      color: white;
      border-radius: 28px;
      padding: 32px;
      box-shadow: var(--shadow);
      position: relative;
      overflow: hidden;
    }}
    .hero::after {{
      content: "";
      position: absolute;
      inset: auto -80px -120px auto;
      width: 260px;
      height: 260px;
      background: rgba(255,255,255,0.08);
      border-radius: 50%;
    }}
    .eyebrow {{
      text-transform: uppercase;
      letter-spacing: 0.18em;
      font-size: 0.78rem;
      opacity: 0.75;
      margin-bottom: 12px;
    }}
    h1 {{ margin: 0; font-size: clamp(2rem, 4vw, 3.2rem); line-height: 1.05; max-width: 780px; }}
    .hero-meta {{ margin-top: 16px; color: rgba(255,255,255,0.84); display: flex; gap: 18px; flex-wrap: wrap; }}
    .hero-meta span {{ background: rgba(255,255,255,0.12); border-radius: 999px; padding: 8px 12px; }}
    nav {{ margin: 18px 0 28px; display: flex; flex-wrap: wrap; gap: 10px; }}
    nav a {{
      text-decoration: none;
      color: var(--accent);
      background: rgba(255,255,255,0.78);
      border: 1px solid rgba(16, 61, 96, 0.08);
      border-radius: 999px;
      padding: 9px 14px;
      font-size: 0.92rem;
      backdrop-filter: blur(6px);
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--panel-border);
      border-radius: 24px;
      padding: 24px;
      margin-bottom: 18px;
      box-shadow: 0 14px 36px rgba(18, 38, 63, 0.07);
    }}
    .panel-head {{ display: flex; align-items: center; justify-content: space-between; gap: 16px; margin-bottom: 16px; }}
    .panel h2 {{ margin: 0; font-size: 1.32rem; color: var(--accent); }}
    .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 14px; }}
    .stat-card {{
      background: linear-gradient(180deg, rgba(16, 61, 96, 0.08), rgba(255,255,255,0.88));
      border: 1px solid rgba(16, 61, 96, 0.08);
      border-radius: 18px;
      padding: 16px;
    }}
    .stat-label {{ font-size: 0.82rem; color: var(--muted); text-transform: uppercase; letter-spacing: 0.08em; }}
    .stat-value {{ margin-top: 8px; font-size: 1.45rem; font-weight: 700; color: var(--accent); word-break: break-word; }}
    .table-wrap {{ overflow-x: auto; border: 1px solid var(--line); border-radius: 18px; }}
    table {{ width: 100%; border-collapse: collapse; background: white; }}
    th, td {{ padding: 12px 14px; border-bottom: 1px solid var(--line); text-align: left; vertical-align: top; font-size: 0.94rem; }}
    th {{ position: sticky; top: 0; background: #133c5d; color: white; font-weight: 600; }}
    tbody tr:nth-child(even) {{ background: #f9fbfd; }}
    tbody tr:hover {{ background: #eef5fa; }}
    .muted {{ color: var(--muted); }}
    .footer {{ text-align: center; color: var(--muted); padding-top: 12px; font-size: 0.9rem; }}
    @media (max-width: 720px) {{
      .shell {{ padding: 20px 14px 40px; }}
      .hero {{ padding: 24px 20px; border-radius: 22px; }}
      .panel {{ padding: 18px; border-radius: 20px; }}
      th, td {{ padding: 10px 11px; font-size: 0.88rem; }}
    }}
  </style>
</head>
<body>
  <div class="shell">
    <header class="hero">
      <div class="eyebrow">OpenStats Export</div>
      <h1>{escape(title)}</h1>
      <div class="hero-meta">
        <span>{escape(organization)}</span>
        <span>{escape(str(dataset.get("name") or "Dataset"))}</span>
        <span>Genere le {escape(_display_datetime(generated_at))}</span>
      </div>
    </header>
    <nav>{nav_links}</nav>
    {rendered_sections}
    <div class="footer">Export dataset {escape(str(dataset.get("id") or ""))} · OpenStats</div>
  </div>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as handle:
        handle.write(html)

    return output_path


def _summary_frame(payload: dict) -> pd.DataFrame:
    metadata = payload.get("metadata", {})
    dataset = payload.get("dataset", {})
    data_summary = payload.get("data_summary", {})
    rows = [
        {"Champ": "Titre", "Valeur": metadata.get("title")},
        {"Champ": "Organisation", "Valeur": metadata.get("organization")},
        {"Champ": "Dataset ID", "Valeur": dataset.get("id")},
        {"Champ": "Nom du dataset", "Valeur": dataset.get("name")},
        {"Champ": "Fichier source", "Valeur": dataset.get("original_filename")},
        {"Champ": "Taille du fichier", "Valeur": _human_size(dataset.get("file_size"))},
        {"Champ": "Cree le", "Valeur": dataset.get("created_at")},
        {"Champ": "Mis a jour le", "Valeur": dataset.get("updated_at")},
        {"Champ": "Lignes (courant)", "Valeur": _dig(dataset, "shape", "rows")},
        {"Champ": "Colonnes (courant)", "Valeur": _dig(dataset, "shape", "columns")},
        {"Champ": "Lignes (brut)", "Valeur": _dig(dataset, "raw_shape", "rows")},
        {"Champ": "Colonnes (brut)", "Valeur": _dig(dataset, "raw_shape", "columns")},
        {"Champ": "Colonnes actives", "Valeur": _dig(dataset, "active_shape", "columns")},
        {"Champ": "Colonnes exclues", "Valeur": ", ".join(dataset.get("excluded_columns") or []) or "Aucune"},
        {"Champ": "Version courante", "Valeur": dataset.get("current_version")},
        {"Champ": "Nombre de versions", "Valeur": dataset.get("versions_count")},
        {"Champ": "Memoire estimee (MB)", "Valeur": data_summary.get("memory_usage_mb")},
        {"Champ": "Export genere le", "Valeur": metadata.get("generated_at")},
    ]
    return pd.DataFrame(rows)


def _preview_frame(payload: dict) -> pd.DataFrame:
    return pd.DataFrame.from_records(payload.get("data_summary", {}).get("preview") or [])


def _dictionary_frame(payload: dict) -> pd.DataFrame:
    rows = []
    for entry in payload.get("data_summary", {}).get("dictionary", []):
        stats = entry.get("stats", {}) or {}
        rows.append({
            "Colonne": entry.get("nom_brut"),
            "Libelle": entry.get("nom_lisible"),
            "Type": entry.get("type_statistique"),
            "Type Regex": entry.get("type_regex"),
            "Unite": entry.get("unite_mesure"),
            "Domaine": entry.get("domaine_unite"),
            "Format Date": entry.get("date_format"),
            "Taux Nullite": entry.get("taux_nullite"),
            "Cardinalite": entry.get("cardinalite"),
            "Moyenne": stats.get("mean"),
            "Mediane": stats.get("median"),
            "Ecart-type": stats.get("std"),
            "Min": stats.get("min"),
            "Max": stats.get("max"),
            "Top Valeurs": _json_text(stats.get("top_values")),
        })
    return pd.DataFrame(rows)


def _descriptive_frames(payload: dict) -> tuple[pd.DataFrame, pd.DataFrame]:
    numeric_rows = []
    categorical_rows = []
    descriptive_stats = payload.get("analysis", {}).get("descriptive_stats") or {}
    for column_name, stats in descriptive_stats.items():
        if not isinstance(stats, dict):
            continue
        if stats.get("type") == "numeric":
            row = {
                "Variable": column_name,
                "Type": stats.get("dtype"),
                "Effectif": stats.get("count"),
                "Moyenne": stats.get("mean"),
                "Mediane": stats.get("median"),
                "Mode": stats.get("mode"),
                "Ecart-type": stats.get("std"),
                "Variance": stats.get("variance"),
                "Min": stats.get("min"),
                "Q1": stats.get("q1"),
                "Q3": stats.get("q3"),
                "Max": stats.get("max"),
                "IQR": stats.get("iqr"),
                "CV (%)": stats.get("cv"),
                "Skewness": stats.get("skewness"),
                "Kurtosis": stats.get("kurtosis"),
                "Nulls": stats.get("null_count"),
                "Taux Nullite": stats.get("null_rate"),
            }
            ci = stats.get("confidence_intervals") or {}
            ci_values = ci.get("bootstrap_ci") or {}
            if ci_values:
                row.update({
                    "IC Moyenne Bas": _dig(ci_values, "mean", "ci_lower"),
                    "IC Moyenne Haut": _dig(ci_values, "mean", "ci_upper"),
                    "IC Mediane Bas": _dig(ci_values, "median", "ci_lower"),
                    "IC Mediane Haut": _dig(ci_values, "median", "ci_upper"),
                    "IC Ecart-type Bas": _dig(ci_values, "std", "ci_lower"),
                    "IC Ecart-type Haut": _dig(ci_values, "std", "ci_upper"),
                    "N Bootstrap": ci.get("n_bootstrap"),
                })
            numeric_rows.append(row)
        else:
            categorical_rows.append({
                "Variable": column_name,
                "Type": stats.get("dtype"),
                "Effectif": stats.get("count"),
                "Cardinalite": stats.get("cardinality"),
                "Mode": stats.get("mode"),
                "Frequence du mode": stats.get("mode_frequency"),
                "Top Valeurs": _json_text(stats.get("top_values")),
                "Nulls": stats.get("null_count"),
                "Taux Nullite": stats.get("null_rate"),
            })
    return pd.DataFrame(numeric_rows), pd.DataFrame(categorical_rows)


def _correlation_matrix_frame(payload: dict, method: str) -> pd.DataFrame:
    corr = payload.get("analysis", {}).get("correlations", {}).get(method) or {}
    matrix = corr.get("matrix")
    columns = corr.get("columns") or []
    if not matrix or not columns:
        return pd.DataFrame()
    frame = pd.DataFrame(matrix)
    return frame.reindex(index=columns, columns=columns).reset_index(names="Variable")


def _significant_correlations_frame(payload: dict) -> pd.DataFrame:
    rows = []
    correlations = payload.get("analysis", {}).get("correlations", {})
    for method, corr in correlations.items():
        for pair in corr.get("significant_pairs", []):
            rows.append({
                "Methode": method,
                "Variable 1": pair.get("var1"),
                "Variable 2": pair.get("var2"),
                "Coefficient": pair.get("coefficient"),
                "Force": pair.get("strength"),
            })
    return pd.DataFrame(rows)


def _vif_frame(payload: dict) -> pd.DataFrame:
    return pd.DataFrame(payload.get("analysis", {}).get("vif") or [])


def _tests_frame(payload: dict) -> pd.DataFrame:
    rows = []
    for test in payload.get("tests") or []:
        rows.append({
            "Test": _coalesce(test.get("test_name"), test.get("test"), test.get("test_type")),
            "Type": test.get("test_type"),
            "Statistique": test.get("statistic"),
            "P-value": test.get("p_value"),
            "Significatif": test.get("significant"),
            "Taille d effet": _json_text(test.get("effect_size")),
            "Interpretation": test.get("interpretation"),
            "Erreur": test.get("error"),
        })
    return pd.DataFrame(rows)


def _model_ranking_frame(payload: dict) -> pd.DataFrame:
    modeling = payload.get("modeling") or {}
    rows = []
    for entry in modeling.get("ranking") or []:
        row = {
            "Rang": entry.get("rank"),
            "Modele": _coalesce(entry.get("model_name"), entry.get("name"), entry.get("model"), entry.get("key")),
            "Cle": _coalesce(entry.get("model_key"), entry.get("key")),
            "Type de tache": _coalesce(entry.get("task_type"), modeling.get("task_type")),
            "CV Mean": _dig(entry, "cv_scores", "mean"),
            "CV Std": _dig(entry, "cv_scores", "std"),
            "Parametres": _json_text(entry.get("best_params")),
        }
        for metric_name, metric_value in (entry.get("metrics") or {}).items():
            clean_metric = _sanitize_for_json(metric_value)
            if isinstance(clean_metric, (str, int, float, bool)) and clean_metric is not None:
                row[_labelize(metric_name)] = clean_metric
        rows.append(row)
    return pd.DataFrame(rows)


def _feature_importance_frame(payload: dict) -> pd.DataFrame:
    rows = []
    for entry in payload.get("modeling", {}).get("ranking") or []:
        model_name = _coalesce(entry.get("model_name"), entry.get("name"), entry.get("key"))
        for feature in entry.get("feature_importance") or []:
            rows.append({
                "Modele": model_name,
                "Feature": feature.get("feature"),
                "Importance": feature.get("importance"),
            })
    return pd.DataFrame(rows)


def _shap_frame(payload: dict) -> pd.DataFrame:
    shap = payload.get("modeling", {}).get("shap") or {}
    return pd.DataFrame(shap.get("global_importance") or [])


def _timeseries_summary_frame(payload: dict) -> pd.DataFrame:
    ts = payload.get("timeseries") or {}
    if not ts:
        return pd.DataFrame()
    return pd.DataFrame([{
        "Variable date": ts.get("date_col"),
        "Variable valeur": ts.get("value_col"),
        "Observations": ts.get("n_observations"),
        "Frequence": ts.get("frequency"),
        "Periode saisonniere": ts.get("seasonal_period"),
        "Meilleur modele": ts.get("best_model"),
        "Plage debut": _dig(ts, "date_range", "start"),
        "Plage fin": _dig(ts, "date_range", "end"),
        "Erreur": ts.get("error"),
    }])


def _timeseries_forecast_frame(payload: dict) -> pd.DataFrame:
    ts = payload.get("timeseries") or {}
    best_model = _select_best_univariate_model(ts)
    if not best_model:
        return pd.DataFrame()

    forecast = best_model.get("forecast") or {}
    dates = forecast.get("dates") or []
    values = forecast.get("values") or []
    lower = forecast.get("lower_ci") or [None] * len(dates)
    upper = forecast.get("upper_ci") or [None] * len(dates)
    rows = []
    for idx, date in enumerate(dates):
        rows.append({
            "Modele": best_model.get("model"),
            "Date": date,
            "Prevision": values[idx] if idx < len(values) else None,
            "IC Bas": lower[idx] if idx < len(lower) else None,
            "IC Haut": upper[idx] if idx < len(upper) else None,
        })
    return pd.DataFrame(rows)


def _multivariate_summary_frame(payload: dict) -> pd.DataFrame:
    mts = payload.get("multivariate_timeseries") or {}
    if not mts:
        return pd.DataFrame()
    orders = _dig(mts, "integration_diagnostics", "orders") or {}
    return pd.DataFrame([{
        "Variable date": mts.get("date_col"),
        "Variables": ", ".join(mts.get("value_cols") or []),
        "Observations": mts.get("n_observations"),
        "Nombre de series": mts.get("n_variables"),
        "Frequence": mts.get("frequency"),
        "Toutes stationnaires": mts.get("all_stationary"),
        "Ordres integration": ", ".join(f"{col}=I({order})" for col, order in orders.items()),
        "Johansen/VECM valide": _dig(mts, "johansen_cointegration", "assumption_valid"),
        "Meilleur modele": mts.get("best_model"),
        "Recommandation": mts.get("recommendation"),
        "Diagnostic integration": _dig(mts, "integration_diagnostics", "interpretation"),
        "Debut": _dig(mts, "date_range", "start"),
        "Fin": _dig(mts, "date_range", "end"),
        "Erreur": mts.get("error"),
    }])


def _granger_frame(payload: dict) -> pd.DataFrame:
    details = _dig(payload, "multivariate_timeseries", "granger_causality", "details") or []
    return pd.DataFrame(details)


def _johansen_frame(payload: dict) -> pd.DataFrame:
    johansen = _dig(payload, "multivariate_timeseries", "johansen_cointegration") or {}
    rows = []
    if johansen:
        rows.append({
            "Type de test": "Synthese",
            "Hypothese": "Hypothèse I(1)",
            "Statistique": johansen.get("assumption_message"),
            "Valeur critique 95%": None,
            "Rejet": johansen.get("assumption_valid"),
        })
    for test in johansen.get("trace_tests") or []:
        rows.append({
            "Type de test": "Trace",
            "Hypothese": test.get("hypothesis"),
            "Statistique": test.get("statistic"),
            "Valeur critique 95%": test.get("critical_value_95"),
            "Rejet": test.get("reject"),
        })
    for test in johansen.get("max_eigenvalue_tests") or []:
        rows.append({
            "Type de test": "Max Eigenvalue",
            "Hypothese": test.get("hypothesis"),
            "Statistique": test.get("statistic"),
            "Valeur critique 95%": test.get("critical_value_95"),
            "Rejet": test.get("reject"),
        })
    return pd.DataFrame(rows)


def _multivariate_forecast_frame(payload: dict) -> pd.DataFrame:
    mts = payload.get("multivariate_timeseries") or {}
    best_key = mts.get("best_model")
    models = mts.get("models") or {}
    best_model = models.get(best_key) if best_key else None
    if not best_model:
        for candidate in models.values():
            if candidate and not candidate.get("error"):
                best_model = candidate
                break
    if not best_model:
        return pd.DataFrame()

    forecast = best_model.get("forecast") or {}
    dates = forecast.get("dates") or []
    series = forecast.get("series") or {}
    if not dates or not series:
        return pd.DataFrame()

    rows = []
    for idx, date in enumerate(dates):
        row = {
            "Modele": best_model.get("model"),
            "Date": date,
        }
        for var_name, values in series.items():
            row[var_name] = values[idx] if idx < len(values) else None
        rows.append(row)
    return pd.DataFrame(rows)


def _pca_variance_frame(payload: dict) -> pd.DataFrame:
    pca = _dig(payload, "factor_analysis", "pca") or {}
    labels = pca.get("component_labels") or []
    if not labels:
        return pd.DataFrame()
    rows = []
    eigenvalues = pca.get("eigenvalues") or []
    explained = pca.get("explained_variance_ratio") or []
    cumulative = pca.get("cumulative_variance") or []
    for idx, label in enumerate(labels):
        rows.append({
            "Composante": label,
            "Valeur propre": eigenvalues[idx] if idx < len(eigenvalues) else None,
            "Variance expliquee": explained[idx] if idx < len(explained) else None,
            "Variance cumulee": cumulative[idx] if idx < len(cumulative) else None,
        })
    return pd.DataFrame(rows)


def _pca_loadings_frame(payload: dict) -> pd.DataFrame:
    pca = _dig(payload, "factor_analysis", "pca") or {}
    loadings = pca.get("loadings") or {}
    rows = []
    for variable, coords in loadings.items():
        row = {"Variable": variable}
        row.update(coords)
        contrib = _dig(pca, "contrib_var", variable) or {}
        for component, value in contrib.items():
            row[f"Contribution {component}"] = value
        rows.append(row)
    return pd.DataFrame(rows)


def _ca_coords_frame(payload: dict, axis: str) -> pd.DataFrame:
    ca = _dig(payload, "factor_analysis", "ca") or {}
    source_key = "row_coords" if axis == "row" else "col_coords"
    contrib_key = "row_contrib" if axis == "row" else "col_contrib"
    cos2_key = "row_cos2" if axis == "row" else "col_cos2"
    coords = ca.get(source_key) or {}
    rows = []
    for label, values in coords.items():
        row = {"Libelle": label}
        row.update(values)
        for component, value in (ca.get(contrib_key, {}).get(label) or {}).items():
            row[f"Contribution {component}"] = value
        for component, value in (ca.get(cos2_key, {}).get(label) or {}).items():
            row[f"Cos2 {component}"] = value
        rows.append(row)
    return pd.DataFrame(rows)


def _mca_modalities_frame(payload: dict) -> pd.DataFrame:
    mca = _dig(payload, "factor_analysis", "mca") or {}
    modality_info = mca.get("modality_info") or []
    coords = mca.get("modality_coords") or {}
    contrib = mca.get("modality_contrib") or {}
    cos2 = mca.get("modality_cos2") or {}
    rows = []
    for info in modality_info:
        full_name = info.get("full")
        row = {
            "Variable": info.get("variable"),
            "Modalite": info.get("modality"),
            "Cle": full_name,
        }
        row.update(coords.get(full_name) or {})
        for component, value in (contrib.get(full_name) or {}).items():
            row[f"Contribution {component}"] = value
        for component, value in (cos2.get(full_name) or {}).items():
            row[f"Cos2 {component}"] = value
        rows.append(row)
    return pd.DataFrame(rows)


def _log_frame(records: list[dict] | None) -> pd.DataFrame:
    return _flatten_records(records or [])


def _versions_frame(payload: dict) -> pd.DataFrame:
    return _flatten_records(payload.get("versions") or [])


def _history_frame(payload: dict) -> pd.DataFrame:
    return _flatten_records(payload.get("history") or [])


def _audit_frame(payload: dict) -> pd.DataFrame:
    return _flatten_records(payload.get("audit_trail") or [])


def _flatten_records(records: list[dict]) -> pd.DataFrame:
    rows = []
    for record in records:
        row = {}
        for key, value in (record or {}).items():
            row[_labelize(key)] = _stringify_nested(value)
        rows.append(row)
    return pd.DataFrame(rows)


def _render_summary_cards(payload: dict) -> str:
    dataset = payload.get("dataset", {})
    metadata = payload.get("metadata", {})
    cards = [
        ("Dataset", dataset.get("name")),
        ("Fichier source", dataset.get("original_filename")),
        ("Lignes", _dig(dataset, "shape", "rows")),
        ("Colonnes", _dig(dataset, "shape", "columns")),
        ("Colonnes actives", _dig(dataset, "active_shape", "columns")),
        ("Version", dataset.get("current_version")),
        ("Export", _display_datetime(metadata.get("generated_at"))),
        ("Colonnes exclues", ", ".join(dataset.get("excluded_columns") or []) or "Aucune"),
    ]
    cards_html = "".join(
        f'''
        <div class="stat-card">
          <div class="stat-label">{escape(str(label))}</div>
          <div class="stat-value">{escape(_fmt(value))}</div>
        </div>
        '''
        for label, value in cards
    )
    return f'<div class="stats-grid">{cards_html}</div>'


def _render_dataframe(frame: pd.DataFrame, max_rows: int = 40) -> str:
    if frame.empty:
        return '<p class="muted">Aucune donnee disponible.</p>'

    clipped = frame.head(max_rows).copy()
    headers = "".join(f"<th>{escape(str(column))}</th>" for column in clipped.columns)
    body_rows = []
    for _, row in clipped.iterrows():
        cells = "".join(f"<td>{escape(_fmt(row[column]))}</td>" for column in clipped.columns)
        body_rows.append(f"<tr>{cells}</tr>")
    notice = ""
    if len(frame) > max_rows:
        notice = f'<p class="muted">Affichage limite a {max_rows} lignes sur {len(frame)}.</p>'
    return (
        f'{notice}<div class="table-wrap"><table><thead><tr>{headers}</tr></thead>'
        f'<tbody>{"".join(body_rows)}</tbody></table></div>'
    )


def _select_best_univariate_model(timeseries: dict) -> dict | None:
    if not timeseries:
        return None
    models = timeseries.get("models") or {}
    best_key = timeseries.get("best_model")
    if best_key and best_key in models and not models[best_key].get("error"):
        return models[best_key]
    for candidate in models.values():
        if candidate and not candidate.get("error"):
            return candidate
    return None


def _safe_sheet_name(base_name: str, used_names: set[str]) -> str:
    clean = base_name[:31]
    suffix = 1
    candidate = clean
    while candidate in used_names:
        suffix_str = f"_{suffix}"
        candidate = f"{clean[:31 - len(suffix_str)]}{suffix_str}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def _coalesce(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return None


def _dig(obj: dict | None, *keys: str) -> Any:
    current = obj or {}
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def _labelize(value: str) -> str:
    return value.replace("_", " ").strip().title()


def _stringify_nested(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return _json_text(value)
    return _sanitize_for_json(value)


def _json_text(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(_sanitize_for_json(value), ensure_ascii=False)


def _display_datetime(value: Any) -> str:
    if not value:
        return ""
    text = str(value)
    if "T" in text:
        return text.replace("T", " ").replace("+00:00", " UTC")
    return text


def _slugify(text: str) -> str:
    return "-".join(part for part in "".join(ch.lower() if ch.isalnum() else " " for ch in text).split() if part)


def _human_size(num_bytes: Any) -> str:
    value = _sanitize_for_json(num_bytes)
    if value in (None, ""):
        return ""
    size = float(value)
    units = ["B", "KB", "MB", "GB"]
    unit_idx = 0
    while size >= 1024 and unit_idx < len(units) - 1:
        size /= 1024
        unit_idx += 1
    return f"{size:.2f} {units[unit_idx]}"


def _fmt(value: Any) -> str:
    value = _sanitize_for_json(value)
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, int):
        return f"{value:,}".replace(",", " ")
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return "—"
        abs_value = abs(value)
        if abs_value == 0:
            return "0"
        if abs_value >= 1000:
            return f"{value:,.2f}".replace(",", " ")
        if abs_value >= 1:
            return f"{value:.4f}"
        if abs_value >= 0.001:
            return f"{value:.6f}"
        return f"{value:.3e}"
    return str(value)


def _sanitize_for_json(obj: Any) -> Any:
    if isinstance(obj, dict):
        return {str(key): _sanitize_for_json(value) for key, value in obj.items()}
    if isinstance(obj, (list, tuple, set)):
        return [_sanitize_for_json(value) for value in obj]
    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()
    if isinstance(obj, pd.Timedelta):
        return str(obj)
    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass
    if hasattr(obj, "item"):
        try:
            return _sanitize_for_json(obj.item())
        except Exception:
            pass
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    return obj
